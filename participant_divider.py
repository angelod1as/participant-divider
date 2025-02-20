# %%
input("Press any key to start...")
print("Starting script...")

# %% [markdown]
# # Separating participants into groups
#
# Simple script that builds a sheet in an excel file.
#
# ## Usage
#
# In a camp participant's spreadsheet, this script takes the paying and confirmed participants and separates them into groups by age.

# %% [markdown]
# Gets the first XLSX in the same folder as the script as main file.

# %%
from pathlib import Path

file_folder = Path("./")


def find_first_xlsx():
    for filename in file_folder.glob("*.xlsx"):
        print(filename)
        if "xlsx" in str(filename):
            return Path(filename)


path = find_first_xlsx()
print(f"\tFound file {path}")

# %% [markdown]
# Turns the sheet into a dataframe and sets the phone as string (to avoid confusion with numbers and zeroes)

# %%
import pandas as pd

df = pd.read_excel(
    path,
    sheet_name="LISTA DE PAGANTES (FINANCEIRA)",
    header=9,
)

# %% [markdown]
# Turns `Desistente` and `Pago` into booleans

# %%
print("\tFiltering out dropouts, non-paid, and team members...")

df["Desistente"] = df["Desistente"] != "NÃO"
df["Pago?"] = df["Pago?"] == "PAGANTE"

# %% [markdown]
# Filters out team members — leaving only participants

# %%
df = df[df["Será Acampante ou Equipe?"] == "Sou participante (acampante)"]

# %% [markdown]
# Filters out non-paying and dropouts

# %%
df = df[df["Pago?"]]
df = df[~df["Desistente"]]

# %% [markdown]
# Remove useless columns

# %%
print("\tRemoving unused columns...")

df.drop(
    columns=[
        'Telefone para contato (Ex: 11XXXXXXXXX sem traço "-" e sem pontos ".")',
        "Pago?",
        "Desistente",
    ],
    inplace=True,
)

# %% [markdown]
# Order by age and gender

# %%
df.sort_values(["Idade"], ascending=[True], inplace=True)

# %% [markdown]
# Create new column and add group using **round-robyn**

# %%
print("\tSeparating into 6 groups...")

df["Grupo"] = 0

for i, index in enumerate(df.index):
    df.loc[index, "Grupo"] = (i % 6) + 1

# %% [markdown]
# Prepare new filename

# %%
import re

print("\tWriting xlsx file...")

pattern = r"([^/\\]*)(?=\.xlsx)"
print(re.search(pattern, str(path)))
match = re.search(pattern, str(path))

if not match:
    raise Exception('Wait, is there an xlsx file inside "source" folder?')

old_name = match.group(0)
new_path = re.sub(pattern, f"{old_name} - COM GRUPOS", str(path))

# %% [markdown]
# Write new excel file

# %%
import openpyxl

workbook = openpyxl.load_workbook(path)

# %% [markdown]
# Add a sheet with every person and their groups

# %%
print("\tCreating unified sheet...")

# FIRST: a unified sheet
unified_sheet_name = "Grupos (todos)"
if unified_sheet_name in workbook.sheetnames:
    unified_sheet = workbook[unified_sheet_name]
else:
    unified_sheet = workbook.create_sheet(unified_sheet_name)

# Clear sheet
unified_sheet.delete_rows(1, unified_sheet.max_row)

row_index = 1

# Write the headers
for c, col_name in enumerate(df.columns):
    cell = unified_sheet.cell(row=1, column=c + 1)
    cell.value = col_name

# Write rows
for r in range(df.shape[0]):
    for c in range(df.shape[1]):
        cell = unified_sheet.cell(row=r + 2, column=c + 1)
        cell.value = df.iloc[r, c]

# Apply AutoFilter
start_column = 1  # Column A (1)
end_column = len(df.columns)  # Number of columns
end_row = df.shape[0] + 1  # Last row including header

filter_range = f"A1:{chr(64 + end_column)}{end_row}"
unified_sheet.auto_filter.ref = filter_range

# %% [markdown]
# Generate a sheet with separated groups

# %%
print("\tCreating separated sheet...")

# THEN, A SEPARATED SHEET
separated_sheet_name = "Grupos (separado)"

if separated_sheet_name in workbook.sheetnames:
    separated_sheet = workbook[separated_sheet_name]
else:
    separated_sheet = workbook.create_sheet(separated_sheet_name)

# Clear sheet
separated_sheet.delete_rows(1, separated_sheet.max_row)

dfs = {group: group_df for group, group_df in df.groupby("Grupo")}

row_index = 1

for group, group_df in dfs.items():
    separated_sheet.cell(row=row_index, column=1, value=f"Grupo {group}")
    row_index += 2  # Move down 2 rows, one for empty space

    # Write headers
    for c, col_name in enumerate(group_df.columns):
        cell = separated_sheet.cell(row=row_index, column=c + 1)
        cell.value = col_name

    row_index += 1  # Move to the next row for data

    # Write data
    for r in range(group_df.shape[0]):
        for c in range(group_df.shape[1]):
            cell = separated_sheet.cell(row=row_index + r, column=c + 1)
            cell.value = group_df.iloc[r, c]

    row_index += group_df.shape[0] + 2  # Move down after each group


# %% [markdown]
# Save the new workbook!!

# %%
workbook.save(new_path)
print(f"{new_path} file written successfully.")
input("Press any key to end...")
