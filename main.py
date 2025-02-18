import re
from pathlib import Path

import openpyxl
import pandas as pd

file_folder = Path("./source")

print("Starting script...")


def find_first_xlsx():
    for filename in file_folder.glob("*.xlsx"):
        if "xlsx" in str(filename):
            return Path(filename)


path = find_first_xlsx()
print(f"\tFound file {path}")

df = pd.read_excel(
    path,
    sheet_name="LISTA DE PAGANTES (FINANCEIRA)",
    header=9,
)

print("\tFiltering out dropouts, non-paid, and team members...")

df["Desistente"] = df["Desistente"] != "NÃO"
df["Pago?"] = df["Pago?"] == "PAGANTE"

df = df[df["Será Acampante ou Equipe?"] == "Sou participante (acampante)"]
df = df[df["Pago?"]]
df = df[~df["Desistente"]]

print("\tRemoving unused columns...")

df.drop(
    columns=[
        'Telefone para contato (Ex: 11XXXXXXXXX sem traço "-" e sem pontos ".")',
        "Pago?",
        "Desistente",
    ],
    inplace=True,
)

df.sort_values(["Idade"], ascending=[True], inplace=True)

df["Grupo"] = 0

print("\tSeparating into 6 groups...")

for i, index in enumerate(df.index):
    df.loc[index, "Grupo"] = (i % 6) + 1


print("\tWriting xlsx file...")
pattern = r"(?<=/)(.*?)(?=\.xlsx)"
match = re.search(pattern, str(path))

if not match:
    raise Exception('Wait, is there an xlsx file inside "source" folder?')

old_name = match.group(0)
new_path = re.sub(pattern, f"{old_name} - COM GRUPOS", str(path))

sheet_name = "Grupos"

workbook = openpyxl.load_workbook(path)

if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    sheet = workbook.create_sheet(sheet_name)

# Clear sheet
sheet.delete_rows(1, sheet.max_row)

row_index = 1

# Write the headers
for c, col_name in enumerate(df.columns):
    cell = sheet.cell(row=1, column=c + 1)
    cell.value = col_name

# Write rows
for r in range(df.shape[0]):
    for c in range(df.shape[1]):
        cell = sheet.cell(row=r + 2, column=c + 1)
        cell.value = df.iloc[r, c]

# Apply AutoFilter
start_column = 1  # Column A (1)
end_column = len(df.columns)  # Number of columns
end_row = df.shape[0] + 1  # Last row including header

filter_range = f"A1:{chr(64 + end_column)}{end_row}"
sheet.auto_filter.ref = filter_range

# Save the modified workbook
workbook.save(new_path)
print(f"{new_path} file written successfully.")
